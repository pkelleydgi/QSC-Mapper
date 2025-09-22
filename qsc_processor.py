import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

def process_qsc_pricing(source_file_path, output_path=None):
    """
    Process QSC pricing source file and apply mappings to template format.
    
    Mappings:
    - SALES PART → MASTERNO
    - LONG DESCRIPTION → DESCRIPTION
    - NET DEALER → STANDARDCOST
    - List Price → MSRP
    - BRAND is always 'QSC'
    - TAXABLE is always 'Y'
    - USETAXFLAG is always 'Y'
    """
    
    if not output_path:
        base_name = os.path.splitext(os.path.basename(source_file_path))[0]
        output_path = f'/mnt/user-data/outputs/QSC_Processed_{base_name}.xlsx'
    
    print(f"Processing QSC source file: {source_file_path}")
    
    # Read the source file
    try:
        # Try reading with different variations of column names
        source_df = pd.read_excel(source_file_path)
        print(f"Source file has {len(source_df)} rows and {len(source_df.columns)} columns")
        print(f"Column names found: {list(source_df.columns)}")
    except Exception as e:
        print(f"Error reading source file: {e}")
        return None
    
    # Create the output dataframe with template structure
    output_data = []
    
    # Find the source columns (case-insensitive matching)
    source_columns_lower = {col.lower().strip(): col for col in source_df.columns}
    
    # Define mapping with variations
    column_mappings = {
        'SALES PART': ['sales part', 'salespart', 'sales_part', 'part number', 'part no', 'partno', 'part'],
        'LONG DESCRIPTION': ['long description', 'longdescription', 'long_description', 'description', 'desc', 'product description'],
        'NET DEALER': ['net dealer', 'netdealer', 'net_dealer', 'dealer price', 'dealer cost', 'net price', 'net cost', 'dealer'],
        'List Price': ['list price', 'listprice', 'list_price', 'msrp', 'retail price', 'retail', 'price', 'list']
    }
    
    # Find actual column names in source
    found_columns = {}
    for target, variations in column_mappings.items():
        for var in variations:
            if var in source_columns_lower:
                found_columns[target] = source_columns_lower[var]
                print(f"Found '{target}' as '{source_columns_lower[var]}'")
                break
    
    # Check if all required columns were found
    missing_columns = []
    for required in column_mappings.keys():
        if required not in found_columns:
            missing_columns.append(required)
    
    if missing_columns:
        print(f"\nWarning: Could not find the following columns: {missing_columns}")
        print("Attempting to proceed with available columns...")
    
    # Process each row
    for index, row in source_df.iterrows():
        output_row = {
            'MASTERNO': row.get(found_columns.get('SALES PART', ''), ''),
            'PARTNO': row.get(found_columns.get('SALES PART', ''), ''),  # Also map to PARTNO
            'DESCRIPTION': row.get(found_columns.get('LONG DESCRIPTION', ''), ''),
            'BRAND': 'QSC',  # Always QSC
            'STANDARDCOST': row.get(found_columns.get('NET DEALER', ''), ''),
            'MSRP': row.get(found_columns.get('List Price', ''), ''),
            'TAXABLE': 'Y',  # Always Y
            'USETAXFLAG': 'Y',  # Always Y
            'CATEGORY1': '',
            'CATEGORY2': '',
            'CATEGORY3': '',
            'WEIGHT': '',
            'HEIGHT': '',
            'WIDTH': '',
            'DEPTH': '',
            'UPC': '',
            'MANUFACTURERNO': '',
            'VENDORPARTNO': '',
            'NOTES': ''
        }
        
        # Clean numeric values (remove $ signs, commas, etc.)
        for price_field in ['STANDARDCOST', 'MSRP']:
            if output_row[price_field]:
                value = str(output_row[price_field])
                # Remove currency symbols and commas
                value = value.replace('$', '').replace(',', '').strip()
                try:
                    output_row[price_field] = float(value) if value else ''
                except ValueError:
                    output_row[price_field] = value
        
        output_data.append(output_row)
    
    # Create output DataFrame
    output_df = pd.DataFrame(output_data)
    
    # Create Excel writer with formatting
    wb = Workbook()
    sheet = wb.active
    sheet.title = "QSC Pricing"
    
    # Write headers
    headers = list(output_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='366092')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Set column widths
    column_widths = {
        'MASTERNO': 20,
        'DESCRIPTION': 50,
        'BRAND': 10,
        'STANDARDCOST': 15,
        'MSRP': 15,
        'TAXABLE': 10,
        'USETAXFLAG': 12,
        'CATEGORY1': 15,
        'CATEGORY2': 15,
        'CATEGORY3': 15,
        'WEIGHT': 10,
        'HEIGHT': 10,
        'WIDTH': 10,
        'DEPTH': 10,
        'UPC': 20,
        'MANUFACTURERNO': 20,
        'VENDORPARTNO': 20,
        'NOTES': 30
    }
    
    for col_idx, header in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(header, 15)
    
    # Write data
    for row_idx, (_, row) in enumerate(output_df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            value = row[header]
            
            # Handle NaN values
            if pd.isna(value):
                value = ''
            
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Format price columns
            if header in ['STANDARDCOST', 'MSRP'] and value != '':
                try:
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                except:
                    pass
    
    # Add borders
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border_style
    
    # Freeze header row
    sheet.freeze_panes = 'A2'
    
    # Save the file
    wb.save(output_path)
    print(f"\n✅ Processed file saved to: {output_path}")
    print(f"   - Total rows processed: {len(output_df)}")
    print(f"   - Brand: QSC (all rows)")
    print(f"   - TAXABLE: Y (all rows)")
    print(f"   - USETAXFLAG: Y (all rows)")
    
    return output_path

# If running as a script
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python qsc_processor.py <source_file_path> [output_path]")
        print("\nThis script processes QSC pricing files with the following mappings:")
        print("  - SALES PART → MASTERNO")
        print("  - LONG DESCRIPTION → DESCRIPTION")
        print("  - NET DEALER → STANDARDCOST")
        print("  - List Price → MSRP")
        print("  - BRAND is always 'QSC'")
        print("  - TAXABLE is always 'Y'")
        print("  - USETAXFLAG is always 'Y'")
        sys.exit(1)
    
    source_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    process_qsc_pricing(source_file, output_file)
